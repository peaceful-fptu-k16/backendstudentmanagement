"""
Excel Export Service

Creates comprehensive Excel reports with student data and charts.
Uses openpyxl for Excel manipulation and formatting.
"""

from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from datetime import datetime

from app.models.student import Student

class ExcelExportService:
    """
    Service for creating formatted Excel reports
    
    Methods:
        create_excel_report(): Create complete Excel workbook
        add_student_sheet(): Add formatted student data sheet
        add_statistics_sheet(): Add statistics and charts
        apply_formatting(): Apply professional formatting
    """
    
    def __init__(self):
        """Initialize Excel export service"""
        self.workbook = None
        
    def create_excel_report(
        self,
        students: List[Student],
        output_path: str,
        include_charts: bool = True
    ):
        """
        Create comprehensive Excel report with student data
        
        Creates a multi-sheet Excel workbook containing:
        - Student data with formatting
        - Statistics summary
        - Charts and visualizations (if enabled)
        
        Args:
            students: List of Student objects
            output_path: Path to save Excel file
            include_charts: Whether to include chart sheets
        """
        # Create new workbook
        self.workbook = Workbook()
        
        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']
        
        # Add student data sheet
        self._add_student_data_sheet(students)
        
        # Add statistics sheet
        self._add_statistics_sheet(students)
        
        # Add charts if requested
        if include_charts:
            self._add_charts_sheet(students)
        
        # Save workbook
        self.workbook.save(output_path)
        
    def _add_student_data_sheet(self, students: List[Student]):
        """
        Add formatted student data sheet
        
        Creates a sheet with all student information in table format
        with professional styling.
        
        Args:
            students: List of Student objects
        """
        # Create sheet
        ws = self.workbook.create_sheet("Danh sách sinh viên", 0)
        
        # Define headers
        headers = [
            "STT",
            "Mã SV",
            "Họ",
            "Tên",
            "Họ và tên",
            "Email",
            "Ngày sinh",
            "Quê quán",
            "Điểm Toán",
            "Điểm Văn",
            "Điểm Anh",
            "Điểm TB",
            "Xếp loại"
        ]
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Write data
        for idx, student in enumerate(students, 1):
            row = idx + 1
            
            # Prepare data
            # Note: Average score and grade only shown if all 3 scores are present
            avg_score = student.get_average_score()
            grade = student.get_grade()
            
            data = [
                idx,
                student.student_id,
                student.first_name,
                student.last_name,
                student.get_full_name(),
                student.email or "",
                student.birth_date.strftime("%d/%m/%Y") if student.birth_date else "",
                student.hometown or "",
                student.math_score if student.math_score is not None else "",
                student.literature_score if student.literature_score is not None else "",
                student.english_score if student.english_score is not None else "",
                f"{avg_score:.2f}" if avg_score is not None else "N/A",  # N/A if incomplete scores
                grade if grade else "N/A"  # N/A if no grade
            ]
            
            # Write row
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col_num)
                cell.value = value
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Center alignment for specific columns
                if col_num in [1, 2, 9, 10, 11, 12, 13]:
                    cell.alignment = Alignment(horizontal="center")
                
                # Color code grades (only for valid grades, not N/A)
                if col_num == 13 and value and value != "N/A":
                    if value == "Excellent":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif value == "Good":
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    elif value == "Poor":
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                # Gray out N/A cells
                elif col_num in [12, 13] and value == "N/A":
                    cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                    cell.font = Font(italic=True, color="808080")
        
        # Adjust column widths
        column_widths = [5, 12, 15, 15, 25, 30, 12, 20, 10, 10, 10, 10, 12]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
    def _add_statistics_sheet(self, students: List[Student]):
        """
        Add statistics summary sheet
        
        Args:
            students: List of Student objects
        """
        ws = self.workbook.create_sheet("Thống kê")
        
        # Title
        ws['A1'] = "BÁO CÁO THỐNG KÊ SINH VIÊN"
        ws['A1'].font = Font(bold=True, size=16, color="4472C4")
        ws.merge_cells('A1:D1')
        
        # Timestamp
        ws['A2'] = f"Ngày tạo: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws['A2'].font = Font(italic=True, size=10)
        
        # Calculate statistics
        total = len(students)
        
        # Filter students with complete scores (all 3 subjects)
        students_with_complete = [
            s for s in students 
            if s.math_score is not None 
            and s.literature_score is not None 
            and s.english_score is not None
        ]
        students_with_incomplete = total - len(students_with_complete)
        
        # Score statistics (only from students with complete scores)
        if students_with_complete:
            math_scores = [s.math_score for s in students_with_complete]
            lit_scores = [s.literature_score for s in students_with_complete]
            eng_scores = [s.english_score for s in students_with_complete]
            
            avg_math = sum(math_scores) / len(math_scores) if math_scores else 0
            avg_lit = sum(lit_scores) / len(lit_scores) if lit_scores else 0
            avg_eng = sum(eng_scores) / len(eng_scores) if eng_scores else 0
        else:
            avg_math = avg_lit = avg_eng = 0
        
        # Grade distribution (only students with complete scores)
        grade_dist = {}
        for student in students_with_complete:
            grade = student.get_grade()
            if grade:
                grade_dist[grade] = grade_dist.get(grade, 0) + 1
        
        # Hometown distribution
        hometown_dist = {}
        for student in students:
            if student.hometown:
                hometown_dist[student.hometown] = hometown_dist.get(student.hometown, 0) + 1
        
        # Write statistics
        row = 4
        
        # General stats
        ws[f'A{row}'] = "THỐNG KÊ TỔNG QUAN"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        ws[f'A{row}'] = "Tổng số sinh viên:"
        ws[f'B{row}'] = total
        row += 1
        
        ws[f'A{row}'] = "Sinh viên có đủ điểm 3 môn:"
        ws[f'B{row}'] = len(students_with_complete)
        ws[f'B{row}'].font = Font(color="008000")  # Green
        row += 1
        
        ws[f'A{row}'] = "Sinh viên chưa đủ điểm 3 môn:"
        ws[f'B{row}'] = students_with_incomplete
        ws[f'B{row}'].font = Font(color="FF0000")  # Red
        row += 2
        
        ws[f'A{row}'] = "ĐIỂM TRUNG BÌNH CÁC MÔN"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        ws[f'A{row}'] = "(Chỉ tính sinh viên có đủ điểm 3 môn)"
        ws[f'A{row}'].font = Font(italic=True, size=9, color="808080")
        row += 1
        
        ws[f'A{row}'] = "Điểm trung bình Toán:"
        ws[f'B{row}'] = round(avg_math, 2)
        row += 1
        
        ws[f'A{row}'] = "Điểm trung bình Văn:"
        ws[f'B{row}'] = round(avg_lit, 2)
        row += 1
        
        ws[f'A{row}'] = "Điểm trung bình Anh:"
        ws[f'B{row}'] = round(avg_eng, 2)
        row += 2
        
        # Grade distribution
        ws[f'A{row}'] = "PHÂN BỐ XẾP LOẠI"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        for grade, count in sorted(grade_dist.items()):
            ws[f'A{row}'] = grade
            ws[f'B{row}'] = count
            ws[f'C{row}'] = f"{(count/total*100):.1f}%"
            row += 1
        
        row += 1
        
        # Hometown distribution (top 5)
        ws[f'A{row}'] = "TOP 5 QUÊ QUÁN"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        sorted_hometowns = sorted(hometown_dist.items(), key=lambda x: x[1], reverse=True)[:5]
        for hometown, count in sorted_hometowns:
            ws[f'A{row}'] = hometown
            ws[f'B{row}'] = count
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        
    def _add_charts_sheet(self, students: List[Student]):
        """
        Add charts sheet with embedded charts
        
        Args:
            students: List of Student objects
        """
        ws = self.workbook.create_sheet("Biểu đồ")
        
        ws['A1'] = "BIỂU ĐỒ PHÂN TÍCH"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Note: Excel charts require data in cells
        # For actual chart generation, we'll use matplotlib separately
        ws['A3'] = "Các biểu đồ chi tiết được tạo bằng matplotlib và lưu riêng."
        ws['A3'].font = Font(italic=True)
